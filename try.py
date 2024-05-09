# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from datetime import date
import pandas as pd
import numpy as np
import os 
import sqlite3 as lite
import pandas as pd
import os
import time
import re




first_new_list = []
data_list = [] #為同一組grouping



def make_dic(title, url, summary):
    
    return {"title" : title , "url" : url , "summary" : summary}


def get_data():
    
        html = '''

        <html>
    <head>
        <title>News Articles</title>

        <script>
        function toggleContent1(index) {
        
            
            var moreContent = document.getElementsByClassName("summary")[index];
            var moreLink = document.getElementsByClassName("more-link")[index];

        
            if (moreContent.style.display === "none") {
            moreContent.style.display = "block";
            moreLink.textContent = "點擊收起";
            } else {
            moreContent.style.display = "none";
            moreLink.textContent = "內容摘要（點擊展開）";
            }
        }
       
        </script>

        <py-script> </py-script>

        <style>
        /* 初始狀態下隱藏摘要内容 */

        #content {
            margin-bottom: 30px;
            margin-top: 30px;
            width: auto;
            display: inline-block;
        }
        .summary {
            display: none;
            margin-bottom: 30px;
            margin-top: 30px;
            width: auto;
            display: inline-block;
        }
        .more-link {
            margin-left: 10px;
        }
        .title,
        .titles {
            height: 15px;
            margin-top: 20px;
            margin-bottom: 10px;
        }
        </style>
    </head>

    '''
    

        #Excel
        path = fr"{os.getcwd()}\adjustments\Grouping.xlsx"
        df = pd.read_excel(path).replace(np.nan, "", regex = True)
        
        wb = load_workbook(path)
        ws = wb.active

    
        #Booming_Application
        Booming_Application = ws.cell(row=1, column=1).value
        Booming_list = list(df.iloc[:,0])
        Booming_list = [item.strip().lower() for item in Booming_list if item != '']
        
        #Mature Application
        Mature_Application = ws.cell(row=1, column=2).value
        Mature_list = list(df.iloc[:,1])
        Mature_list = [item.strip().lower() for item in Mature_list if item != '']

        #Foundry Related
        foundry_list=[]
        Foundry = ws.cell(row=1, column=3).value
        Foundry_list = list(df.iloc[:,2])
        Foundry_list = [item.strip().lower() for item in Foundry_list if item != '']
    
        #Other
        
        
        
        
        #database
        conn = lite.connect(fr"{os.getcwd()}\databases\2024-04-23_crawled_sorted.db")
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cur.fetchall()
        
        df_table = pd.DataFrame()
        df = pd.DataFrame()
        
        for table in tables:
            if table[0].strip().lower() in Booming_list:
                
                booming = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)#中文的 table 表名，須加上反引號
                booming_table = pd.concat([df_table, booming])
                
                boom_titles = booming_table["title"].astype(str).values
                boom_urls = booming_table["url"].astype(str).values
                boom_sums = booming_table["summary"].astype(str).values
                
                for title, url, summary in zip(boom_titles, boom_urls, boom_sums):
                    
                    dic = make_dic(title, url, summary)
                    first_new_list.append(dic)
                    data_list.append(dic)
                    
                if data_list[0]["title"] == first_new_list[0]["title"]:
                    data_list.remove(data_list[0])
                
                for i, data in enumerate(first_new_list):
        
                    html_boom = fr'''
                    
                    <body>
                        <h3><b>The Analyzed Htmls of Automatic Semiconductor News.</b></h3>
                        
                        <h2> Booming_Application: </h2>
                        <div class="title">
                        <a
                            href= {first_new_list[0]["url"]}
                        >
                            <b
                            ><font size="4"
                                >{(1)} Title: {first_new_list[0]["title"]}</font
                            ></b
                            >
                        </a>
                        </div>
                        <!-- 摘要内容 -->
                        <div id="content">
                        {first_new_list[0]["summary"]}
                        </div>
                    '''

                n = 2
                html_boom_hidden =''
                for i, data in enumerate(data_list):
                    
                    html_boom_hidden += f'''
                    
                        <div class="titles">
                        <a
                            href= {data["url"]}                           
                        >
                            <b
                            ><font size="4"
                                >{(n)} Title:{data["title"]}</font
                            ></b
                            >
                        </a>

                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="more-link" onclick="toggleContent({i}); return false;"
                            >內容摘要（點擊展開）</a
                        >
                        </div>

                        <!-- 摘要内容 -->
                        <div class="summary">
                        {data["summary"]}
                        </div>

                    </body>
                    </html>

                    '''
                    n+=1
                
                html += html_boom + html_boom_hidden
                
                
                
            elif table[0].strip().lower() in Mature_list:
               
                mature = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)
                mature_table = pd.concat([df_table, mature])

                mature_titles = mature_table["title"].astype(str).values
                mature_urls = mature_table["url"].astype(str).values
                mature_sums = mature_table["summary"].astype(str).values
                    
                for title, url, summary in zip(mature_titles, mature_urls, mature_sums):

                    dic = make_dic(title, url, summary)
                    first_new_list.append(dic)
                    data_list.append(dic)
                
                if data_list[0]["title"] == first_new_list[0]["title"]:
                    data_list.remove(data_list[0])

                
                for i, data in enumerate(first_new_list):
        
                    html_mature = fr'''
                    
                    <body>
                        
                        <h2> {Mature_Application}: </h2>
                        <div class="title">
                        <a
                            href= {first_new_list[0]["url"]}
                        >
                            <b
                            ><font size="4"
                                >{(1)} Title: {first_new_list[0]["title"]}</font
                            ></b
                            >
                        </a>
                        </div>
                        <!-- 摘要内容 -->
                        <div id="content">
                        {first_new_list[0]["summary"]}
                        </div>
                    '''

                n = 2
                html_mature_hidden =''
                for i, data in enumerate(data_list):
                    
                    html_mature_hidden += f'''
                    
                        <div class="titles">
                        <a
                            href= {data["url"]}                           
                        >
                            <b
                            ><font size="4"
                                >{(n)} Title:{data["title"]}</font
                            ></b
                            >
                        </a>

                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="more-link" onclick="toggleContent({i}); return false;"
                            >內容摘要（點擊展開）</a
                        >
                        </div>

                        <!-- 摘要内容 -->
                        <div class="summary">
                        {data["summary"]}
                        </div>

                    </body>
                    </html>

                    '''
                    n+=1
                
                html += html_mature + html_mature_hidden
                
                
            elif table[0].strip().lower() in Foundry_list:
                print(table[0])

            

        return html










if __name__ == '__main__':
    
    html = get_data()
    with open (fr"{date.today()}_test.html", "w", encoding="utf-8") as f:
        f.write (html)
        
    