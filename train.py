from openpyxl import load_workbook
from datetime import date

import sqlite3 as lite
import pandas as pd
import numpy as np
import os 


class Generate_html():
    
    def __init__(self, producible):
        
        self.producible = producible
        print("Start Generating")

    def make_dic(self, title, url, summary):
        
        return {"title" : title , "url" : url , "summary" : summary}


    def process_firstnews(self, data, application):
        
        html = ''
        for i, entry in enumerate(data):
            html += f'''
                <body>
                    <div class="title">
                    <a
                        href= {data[0]["url"]}
                    >
                        <b
                        ><font size="4"
                            >{(1)}. {data[0]["title"]}</font
                        ></b
                        >
                    </a>
                    </div>
                    <!-- 摘要内容 -->
                    <div id="content">
                    {data[0]["summary"]}
                    </div>
                '''
                
            return f'''
            <body>
                <h2>{application}</h2>
                {html}
            </body>
        '''



    def process_data(self, data, application):
        
        html = ''
        if application == "Booming Application":
            
            for i, entry in enumerate(data):
                html += f'''
                    <div class="titles">
                        <a href="{entry["url"]}">
                            <b><font size="4">{(i+2)}. {entry["title"]}</font></b>
                        </a>
                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="more-link" onclick="toggleContent({i}); return false;">
                            內容摘要（點擊展開）
                        </a>
                    </div>
                    <!-- 摘要内容 -->
                    <div class="summary">
                        {entry["summary"]}
                    </div>
                    
                '''
        elif application == "Mature Application":
            
            for i, entry in enumerate(data):
                html += f'''
                    <div class="titles">
                        <a href="{entry["url"]}">
                            <b><font size="4">{(i+2)}. {entry["title"]}</font></b>
                        </a>
                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="mature-more-link" onclick="toggleMatureContent({i}); return false;">
                            內容摘要（點擊展開）
                        </a>
                    </div>
                    <!-- 摘要内容 -->
                    <div class="mature-summary">
                        {entry["summary"]}
                    </div>
                    
                '''
        elif application == "Foundry Related":
            
            for i, entry in enumerate(data):
                html += f'''
                    <div class="titles">
                        <a href="{entry["url"]}">
                            <b><font size="4">{(i+2)}. {entry["title"]}</font></b>
                        </a>
                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="foundry-more-link" onclick="togglefoundryContent({i}); return false;">
                            內容摘要（點擊展開）
                        </a>
                    </div>
                    <!-- 摘要内容 -->
                    <div class="foundry-summary">
                        {entry["summary"]}
                    </div>
                    
                '''
        elif application == "Others":
            
            for i, entry in enumerate(data):
                html += f'''
                    <div class="titles">
                        <a href="{entry["url"]}">
                            <b><font size="4">{(i+2)}. {entry["title"]}</font></b>
                        </a>
                        <!-- 顯示更多内容的連結 -->
                        <a href="#" class="other-more-link" onclick="toggleotherContent({i}); return false;">
                            內容摘要（點擊展開）
                        </a>
                    </div>
                    <!-- 摘要内容 -->
                    <div class="other-summary">
                        {entry["summary"]}
                    </div>
                    
                '''
            
        return html





    def get_data(self):
        
        current_date = date.today()
        current_date = current_date.strftime("%y%m%d")
        
        # if self.producible:
        #     if len(os.listdir(self.directory)) == 0:
        #         print("no data")
        #         self.msg.attach(MIMEText("There is no new analyzed data/html.", "plain", "utf-8"))
        
        html = '''
            <html>
                <head>
                    <title>News Articles</title>
                    <script>
                        //載入頁面時預設摘要是隱藏的
                        window.onload = function() {
                            var summaryElements = document.getElementsByClassName("summary");
                            var matureSummarys = document.getElementsByClassName("mature-summary")
                            var foundrySummarys = document.getElementsByClassName("foundry-summary")
                            var otherSummarys = document.getElementsByClassName("other-summary")
                            
                            for (var i = 0; i < summaryElements.length; i++) {
                                summaryElements[i].style.display = "none";  
                            }
                            for (var i = 0; i < matureSummarys.length; i++){
                                matureSummarys[i].style.display = "none";
                            }
                            for (var i = 0; i < foundrySummarys.length; i++ ){
                                foundrySummarys[i].style.display = "none";
                            }
                            for (var i = 0; i < otherSummarys.length; i++){
                                otherSummarys[i].style.display = "none";
                            }
                        };

                        function toggleContent(index) {
                            var moreContent = document.getElementsByClassName("summary")[index];
                            var moreLink = document.getElementsByClassName("more-link")[index];
                            if (moreContent.style.display === "none") {
                                moreContent.style.display = "block";
                                moreLink.textContent = "點擊收起";
                            } else {
                                moreContent.style.display = "none";
                                moreLink.textContent = "內容摘要（點擊展開）";
                            }
                        }
                        function toggleMatureContent(index) {
                            var moreContent = document.getElementsByClassName("mature-summary")[index];
                            var moreLink = document.getElementsByClassName("mature-more-link")[index];
                            if (moreContent.style.display === "none") {
                                moreContent.style.display = "block";
                                moreLink.textContent = "點擊收起";
                            } else {
                                moreContent.style.display = "none";
                                moreLink.textContent = "內容摘要（點擊展開）";
                            }
                        }
                        function togglefoundryContent(index) {
                            var moreContent = document.getElementsByClassName("foundry-summary")[index];
                            var moreLink = document.getElementsByClassName("foundry-more-link")[index];
                            if (moreContent.style.display === "none") {
                                moreContent.style.display = "block";
                                moreLink.textContent = "點擊收起";
                            } else {
                                moreContent.style.display = "none";
                                moreLink.textContent = "內容摘要（點擊展開）";
                            }
                        }
                        function toggleotherContent(index) {
                            var moreContent = document.getElementsByClassName("other-summary")[index];
                            var moreLink = document.getElementsByClassName("other-more-link")[index];
                            if (moreContent.style.display === "none") {
                                moreContent.style.display = "block";
                                moreLink.textContent = "點擊收起";
                            } else {
                                moreContent.style.display = "none";
                                moreLink.textContent = "內容摘要（點擊展開）";
                            }
                        }
                    </script>
                    <style>
                        /* 初始狀態下隱藏摘要内容 */
                        
                        body{
                            background-image : url('Daily News Templete.png');
                            background-size: 100%, 100%;
                          
                            background-position: top;
                            background-repeat: no-repeat; 
                        }
                        
                        .theme{
                            font-size: 50px;
                            text-align: center;
                            margin-top: 120px;
                            color:darkblue; 
                        }
                        
                        #content {
                            margin-bottom: 20px;
                            margin-top: 20px;
                            width: auto;
                            display: inline-block;
                        }
                        .summary,
                        .mature-summary,
                        .foundry-summary,
                        .other-summary{
                            display: none;
                            width: auto;
                            display: inline-block;
                            margin-bottom: 10px;
                            margin-top: 10px;
                        }
                        .more-link,
                        .mature-more-link,
                        .foundry-more-link
                        .other-more-link {
                            margin-left: 10px;
                        }
                        div {
                            margin: 0;
                            padding: 0;
                        }
                       
                       
                    </style>
                </head>
        '''
        html_title =f'''<h1 class="theme"><b>{current_date}_Supply Chain Daily News </b></h1><br>'''
        
        html += html_title

        #Excel
        path = fr"{os.getcwd()}\adjustments\Grouping.xlsx"
        df = pd.read_excel(path).replace(np.nan, "", regex = True)
        
        wb = load_workbook(path)
        ws = wb.active

        
        #Booming_Application
        Booming_Application = ws.cell(row=1, column=1).value.strip()
        Booming_list = list(df.iloc[:,0])
        Booming_list = [item.strip().lower() for item in Booming_list if item != '']
        
        #Mature Application
        Mature_Application = ws.cell(row=1, column=2).value.strip()
        Mature_list = list(df.iloc[:,1])
        Mature_list = [item.strip().lower() for item in Mature_list if item != '']

        #Foundry Related
        Foundry = ws.cell(row=1, column=3).value.strip()
        Foundry_list = list(df.iloc[:,2])
        Foundry_list = [item.strip().lower() for item in Foundry_list if item != '']

        #Other
        Other = ws.cell(row=1, column=4).value.strip()
        
            
        #database
        conn = lite.connect(fr"{os.getcwd()}\databases\2024-05-06_crawled_sorted.db")
        cur = conn.cursor()
        cur.execute("SELECT name FROM sqlite_master WHERE type='table';")
        tables = cur.fetchall()
        
        self.df_boom = pd.DataFrame()
        self.df_mature = pd.DataFrame()
        self.df_foundry = pd.DataFrame()
        
        
        combined_data, topical_titles = [], []
        foundry_set, mature_set, boom_set, other_set = set(), set(), set(), set()
        
        for table in tables:
            
            # Booming Application
            if table[0].strip().lower() in Booming_list:
                
                booming = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)
                self.df_boom = pd.concat([self.df_boom, booming], ignore_index=True)
   
            # Mature Application
            elif table[0].strip().lower() in Mature_list:
                
                mature = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)
                self.df_mature = pd.concat([self.df_mature, mature], ignore_index=True)
              
            #Foundry
            elif table[0].strip().lower() in Foundry_list:
                
                #foundry_titles = foundry["title"].astype(str).values
                foundry = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)
                self.df_foundry = pd.concat([self.df_foundry, foundry], ignore_index=True)
                
            #other
            else:
                other = pd.read_sql_query(f"SELECT * FROM `{table[0]}`", conn)
                for row in other.to_dict(orient='records'):
                    title = row["title"]
                    url = row["url"]
                    summary = row["summary"]
                    
                    
                    if title not in other_set:
                        combined_data.append(row)
                        other_set.add(title)
                        other_set.add(url)
                        other_set.add(summary)


        #分類比較：假如 foundry 要跟 boom 和 mature比較 , 另外在和other比較（other有可能帶有foundry的新聞內容）

        try:
            #Booming Application
            boom_first_new, boom_new = [],[]
            
            titles_boom = self.df_boom["title"].astype(str).values
            urls_boom = self.df_boom["url"].astype(str).values
            sums_boom = self.df_boom["summary"].astype(str).values

            for title, url, summary in zip(titles_boom, urls_boom, sums_boom):
                if title not in boom_set:
                    boom_set.add(title)
                    boom_set.add(url)
                    boom_set.add(summary)
                    boom_dic = self.make_dic(title, url, summary)

                    boom_first_new.append(boom_dic)
                    boom_new.append(boom_dic)

            if boom_new[0]["title"] == boom_first_new[0]["title"]:
                boom_new.remove(boom_new[0])

            html += self.process_firstnews(boom_first_new, Booming_Application)
            html += self.process_data(boom_new, Booming_Application)


            #Mature Application
            mature_first_new, mature_new, filter_mature = [], [], []
            titles_mature = self.df_mature["title"].astype(str).values
            urls_mature = self.df_mature["url"].astype(str).values
            sums_mature = self.df_mature["summary"].astype(str).values
            filter_mature.extend(titles_boom)
            

            for title, url, summary in zip(titles_mature, urls_mature, sums_mature):
                if title not in mature_set:
                    mature_set.add(title)
                    mature_set.add(url)
                    mature_set.add(summary)
                    mature_dic = self.make_dic(title, url, summary)

                    mature_first_new.append(mature_dic)
                    mature_new.append(mature_dic)
                    
            to_remove = []
            for i ,item in enumerate(mature_new):
                if mature_new[i]["title"] in filter_mature:
                    to_remove.append(i)
            
            for i in reversed(to_remove):
                del mature_new[i]
                del mature_first_new[i]       
    
            if mature_new[0]["title"] == mature_first_new[0]["title"]:
                mature_new.remove(mature_new[0])

            html += self.process_firstnews(mature_first_new, Mature_Application)
            html += self.process_data(mature_new, Mature_Application)




            #Foundry
            foundry_first_new, foundry_new, filter_foundry = [], [], []
        
            titles_foundry = self.df_foundry["title"].astype(str).values
            urls_foundry = self.df_foundry["url"].astype(str).values
            sums_foundry = self.df_foundry["summary"].astype(str).values
            
            filter_foundry.extend(titles_boom)
            filter_foundry.extend(titles_mature)
            

            for title, url, summary in zip(titles_foundry, urls_foundry, sums_foundry):
                if title not in foundry_set:
                    foundry_set.add(title)
                    foundry_set.add(url)
                    foundry_set.add(summary)
                    foundry_dic = self.make_dic(title, url, summary)
                
                    foundry_first_new.append(foundry_dic)
                    foundry_new.append(foundry_dic)

            to_remove = []
            for i ,item in enumerate(foundry_new):
                if foundry_new[i]["title"] in filter_foundry:
                    to_remove.append(i)
                
            for i in reversed(to_remove): # 5, 4, 3, 2, 1
                del foundry_new[i]
                del foundry_first_new[i]
                
            if foundry_new[0]["title"] == foundry_first_new[0]["title"]:
                foundry_new.remove(foundry_new[0])
            

            html+=self.process_firstnews(foundry_first_new, Foundry)
            html+=self.process_data(foundry_new, Foundry)



            #other
            first_new, data_new = [],[]
            
            topical_titles.extend(titles_boom)
            topical_titles.extend(titles_mature)
            topical_titles.extend(titles_foundry)

            
            filtered_combined_data = [row for row in combined_data if row["title"] not in topical_titles]
            
            for row in filtered_combined_data:
                combined_dic = self.make_dic(row["title"], row["url"], row["summary"])
                first_new.append(combined_dic)
                data_new.append(combined_dic)
            
            
            if data_new[0]["title"] == first_new[0]["title"]:
                data_new.remove(data_new[0])
                
                    
            html+=self.process_firstnews(first_new, Other)
            html+=self.process_data(data_new, Other)
            

            html += '''
                </html>
            '''
            html = html.encode('utf-8')
            
            
            
            #Generate html
            with open(fr"{os.getcwd()}\{current_date}_Supply Chain Daily News.html", "wb") as f:
                f.write(html)
        
        
        
        except Exception as e:
            print(e)
            print("Generate html is failed...Please try again.")
            
            return '''
            <body>
                <h2>There is no new analyzed data:</h2>
            </body>
        '''
        






if __name__ == '__main__':
    
    ge_html = Generate_html(True)
    ge_html.get_data()

# my_list = ["Apple", "Pieapple", "watermelans"]

# for i, item in enumerate(my_list):
#     if my_list[i] == "Apple":
#         del my_list[i]
# print(my_list)

# for table in tables:
#     if table in Booming:
#         #print(table)
#         print(group_df.iloc[0,0])



